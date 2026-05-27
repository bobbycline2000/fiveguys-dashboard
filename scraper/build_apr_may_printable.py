#!/usr/bin/env python3
"""
Build a clean printable April + May 2026 secret shop summary for Store 2065.
Output: data/Shop_AprMay_2026_Printable.xlsx
"""
from __future__ import annotations
import json
from pathlib import Path
from datetime import date as dt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SHOPS_JSON = sorted(
    (ROOT / "data" / "raw" / "marketforce" / "2065").glob("*/shops.json")
)[-1]
PART_JSON = ROOT / "data" / "raw" / "marketforce" / "2065" / "participation.json"
NAME_MAP_JSON = ROOT / "data" / "employee_name_map.json"
OUT = ROOT / "data" / "Shop_AprMay_2026_Printable.xlsx"

# --- Colors ---
RED      = "C8102E"
WHITE    = "FFFFFF"
OFF_WHITE= "F9F9F9"
DARK     = "1A1A1A"
DARK_GREY= "595959"
MID_GREY = "D0D0D0"
LIGHT_GREY="F2F2F2"
NAVY     = "1A2A4A"
GOLD     = "D4A017"
LIGHT_GOLD="FFFBF0"
GREEN    = "2E7D32"
LIGHT_GREEN="E8F5E9"
RED_LIGHT= "FDECEA"
HEADER_BG= "3B1F1F"


def _b(c=MID_GREY):
    s = Side(border_style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _f(c): return PatternFill("solid", start_color=c, end_color=c)

def _set(cell, value=None, bold=False, size=10, fg=DARK, bg=None,
         h="center", v="center", wrap=False, italic=False, border=True, num_fmt=None):
    if value is not None: cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    if bg: cell.fill = _f(bg)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border: cell.border = _b()
    if num_fmt: cell.number_format = num_fmt


def score_bg(p):
    if p is None: return OFF_WHITE
    if p >= 100: return LIGHT_GREEN
    if p >= 90:  return "EFF8E8"
    if p >= 80:  return LIGHT_GOLD
    return RED_LIGHT

def score_fg(p):
    if p is None: return DARK
    if p >= 100: return GREEN
    if p >= 90:  return "4A7A0F"
    if p >= 80:  return GOLD
    return RED


def resolve_full_names(first_names: list[str], name_map: dict) -> list[str]:
    out = []
    for fn in first_names:
        resolved = name_map.get(fn) or name_map.get(fn.capitalize())
        if resolved and "UNRESOLVED" not in resolved:
            out.append(resolved)
        else:
            out.append(fn)  # fallback to first name if unresolved
    return sorted(out)


def build():
    shops_data = json.loads(SHOPS_JSON.read_text(encoding="utf-8"))
    all_shops = sorted(shops_data["shops"], key=lambda s: s["date"])
    by_shop = {}
    if PART_JSON.exists():
        by_shop = json.loads(PART_JSON.read_text(encoding="utf-8")).get("by_shop", {})
    name_map = {}
    if NAME_MAP_JSON.exists():
        name_map = json.loads(NAME_MAP_JSON.read_text(encoding="utf-8"))

    # Filter April + May 2026
    apr_may = [s for s in all_shops
               if s.get("date", "").startswith("2026-04") or s.get("date", "").startswith("2026-05")]

    wb = Workbook()
    ws = wb.active
    ws.title = "Apr-May 2026 Summary"
    ws.sheet_view.showGridLines = False

    # Page setup — print-friendly, letter, landscape, fit to 1 page wide
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Column widths
    COL_W = {"A": 6, "B": 10, "C": 9, "D": 9, "E": 9, "F": 52}
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    r = 1

    # === TITLE ===
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws.cell(r, 1),
         "SECRET SHOP SUMMARY — APRIL & MAY 2026",
         bold=True, size=16, fg=WHITE, bg=RED, border=False, h="center")
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"A{r}:F{r}")
    _set(ws.cell(r, 1),
         "Five Guys Store 2065  |  Louisville, KY  |  Printed: May 27, 2026  |  Source: Marketforce / KnowledgeForce",
         size=8, fg=DARK_GREY, bg=LIGHT_GREY, italic=True, border=False, h="center")
    ws.row_dimensions[r].height = 13
    r += 1

    ws.row_dimensions[r].height = 6; r += 1  # spacer

    # === SUMMARY ROW ===
    apr_shops = [s for s in apr_may if s["date"].startswith("2026-04")]
    may_shops = [s for s in apr_may if s["date"].startswith("2026-05")]

    def month_avg(shops_list):
        scored = [s for s in shops_list if s.get("score") is not None]
        if not scored: return None
        return round(sum(s["score"] for s in scored) / len(scored), 1)

    apr_avg = month_avg(apr_shops)
    may_avg = month_avg(may_shops)
    apr_100 = sum(1 for s in apr_shops if s.get("score") == 100)
    may_100 = sum(1 for s in may_shops if s.get("score") == 100)

    # Summary header
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws.cell(r, 1), "PERIOD SUMMARY", bold=True, size=10, fg=WHITE, bg=NAVY, h="left", border=False)
    ws.row_dimensions[r].height = 18; r += 1

    # Summary columns
    for col in range(1, 7):
        _set(ws.cell(r, col), bg=LIGHT_GREY, border=True)
    _set(ws.cell(r, 1), "Period", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    _set(ws.cell(r, 2), "Shops", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    _set(ws.cell(r, 3), "Avg Score", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    _set(ws.cell(r, 4), "High", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    _set(ws.cell(r, 5), "Low", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    _set(ws.cell(r, 6), "100% Shops (Payout)", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="center")
    ws.row_dimensions[r].height = 16; r += 1

    def row_for_month(label, shops_list, avg, cnt100):
        scored = [s["score"] for s in shops_list if s.get("score") is not None]
        high = max(scored) if scored else None
        low  = min(scored) if scored else None
        avg_str = f"{avg}%" if avg is not None else "—"
        high_str = f"{high}%" if high is not None else "—"
        low_str  = f"{low}%" if low  is not None else "—"
        bg_avg = score_bg(avg) if avg else OFF_WHITE
        fg_avg = score_fg(avg) if avg else DARK
        return label, str(len(shops_list)), avg_str, high_str, low_str, f"{cnt100} shop{'s' if cnt100 != 1 else ''}"

    for label, shops_list, avg, cnt100 in [
        ("April 2026", apr_shops, apr_avg, apr_100),
        ("May 2026",   may_shops, may_avg, may_100),
    ]:
        cols_data = row_for_month(label, shops_list, avg, cnt100)
        for ci, val in enumerate(cols_data, 1):
            _set(ws.cell(r, ci), val, size=10, fg=DARK, bg=OFF_WHITE, h="center")
        # color the avg cell
        avg_val = apr_avg if label.startswith("April") else may_avg
        ws.cell(r, 3).fill = _f(score_bg(avg_val))
        ws.cell(r, 3).font = Font(name="Arial", size=10, bold=True, color=score_fg(avg_val))
        ws.row_dimensions[r].height = 16; r += 1

    ws.row_dimensions[r].height = 6; r += 1  # spacer

    # === SHOP-BY-SHOP DETAIL ===
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws.cell(r, 1), "SHOP-BY-SHOP DETAIL", bold=True, size=10, fg=WHITE, bg=HEADER_BG, h="left", border=False)
    ws.row_dimensions[r].height = 18; r += 1

    # Column headers
    headers = ["#", "Date", "Meal Period", "Score", "Payout", "Employees on Shift"]
    for ci, h_text in enumerate(headers, 1):
        _set(ws.cell(r, ci), h_text, bold=True, size=9, fg=WHITE, bg=NAVY, h="center")
    ws.row_dimensions[r].height = 16; r += 1

    for idx, shop in enumerate(apr_may, 1):
        shop_date = dt.fromisoformat(shop["date"])
        date_str  = shop_date.strftime("%a %m/%d/%y")
        meal      = shop.get("meal_period", "")
        score     = shop.get("score")
        score_str = f"{score}%" if score is not None else "—"
        payout    = "YES — $50/ea" if score == 100 else "—"

        first_names = by_shop.get(shop["job_id"], [])
        full_names  = resolve_full_names(first_names, name_map)
        names_str   = ", ".join(full_names) if full_names else "—"

        bg_row = LIGHT_GREY if idx % 2 == 0 else WHITE
        bg_score = score_bg(score)
        fg_score = score_fg(score)

        _set(ws.cell(r, 1), idx, size=9, fg=DARK_GREY, bg=bg_row, h="center")
        _set(ws.cell(r, 2), date_str, size=10, fg=DARK, bg=bg_row, h="center", bold=True)
        meal_bg = LIGHT_GOLD if "lunch" in meal.lower() else "E8EEF8"
        meal_fg = GOLD      if "lunch" in meal.lower() else NAVY
        _set(ws.cell(r, 3), meal, size=10, fg=meal_fg, bg=meal_bg, h="center", bold=True)
        _set(ws.cell(r, 4), score_str, size=12, fg=fg_score, bg=bg_score, h="center", bold=True)
        if score == 100:
            _set(ws.cell(r, 5), payout, size=9, fg=GREEN, bg=LIGHT_GREEN, h="center", bold=True)
        else:
            _set(ws.cell(r, 5), payout, size=9, fg=DARK_GREY, bg=bg_row, h="center")
        _set(ws.cell(r, 6), names_str, size=9, fg=DARK, bg=bg_row, h="left", wrap=True)
        # Taller row for names
        n_names = len(full_names)
        ws.row_dimensions[r].height = max(18, 14 * ((n_names // 4) + 1))
        r += 1

    ws.row_dimensions[r].height = 6; r += 1  # spacer

    # === FOOTER ===
    ws.merge_cells(f"A{r}:F{r}")
    _set(ws.cell(r, 1),
         "Payout = $50 per employee who worked during a 100% shop's meal window. "
         "Employees listed are those clocked in during the shop's meal period per CrunchTime CETD + Marketforce participation data.",
         size=7, fg=DARK_GREY, bg=LIGHT_GREY, italic=True, h="left", border=False, wrap=True)
    ws.row_dimensions[r].height = 20

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"April shops: {len(apr_shops)} | May shops: {len(may_shops)} | Total: {len(apr_may)}")
    print(f"April avg: {apr_avg}% | May avg: {may_avg}%")


if __name__ == "__main__":
    build()
