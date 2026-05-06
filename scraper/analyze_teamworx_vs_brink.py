"""
Compare Teamworx forecast hourly Sales vs Par Brink actual hourly Sales for a store.
Builds: BobbyWorkspace/projects/scg/case-studies/teamworx-forecast-accuracy/
        Teamworx_Forecast_vs_Brink_Actual_{store}.xlsx

Reads:
  - {repo}/data/teamworx_hourly_history.json — captured Forecast Graph hourly data
    keyed by date (YYYY-MM-DD) with sales/ideal/scheduled arrays for hours 6AM-11PM
  - {repo}/data/raw/parbrink/{store}/{date}/Hourly Sales And Labor.pdf for each date
    (parses on the fly using the existing parbrink_parse_hourly_sales_labor.py logic)

Usage:
  python scraper/analyze_teamworx_vs_brink.py --store 2065
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = Path("C:/Users/bobby/OneDrive/BobbyWorkspace")
CASE_STUDY = WORKSPACE / "projects" / "scg" / "case-studies" / "teamworx-forecast-accuracy"

LABELS = ["6AM","7AM","8AM","9AM","10AM","11AM","12PM","1PM","2PM","3PM","4PM","5PM","6PM","7PM","8PM","9PM","10PM","11PM"]
HOUR_24 = [6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]


def parse_brink_pdf(pdf_path: Path) -> dict | None:
    """Parse a Brink Hourly Sales And Labor PDF — returns {hour_24: {netSales, guests, laborHrs, laborDollars, laborPct}}"""
    try:
        from pypdf import PdfReader
    except ImportError:
        return None
    if not pdf_path.exists():
        return None
    reader = PdfReader(str(pdf_path))
    text = "\n".join(p.extract_text() or "" for p in reader.pages)
    rows = {}
    re_row = re.compile(
        r"(\d{1,2}:\d{2}\s*[AP]M)\s+\$?([\d,]+\.\d{2})\s+(\d+)\s+\$?[\d,]+\.\d{2}\s+(\d+)\s+\$?[\d,]+\.\d{2}\s+([\d,]+\.\d{2})\s+\$?([\d,]+\.\d{2})\s+([\d,]+\.\d{2})%"
    )
    for m in re_row.finditer(text):
        hour_str = m.group(1).strip()
        h = int(hour_str.split(":")[0])
        if "PM" in hour_str and h != 12:
            h += 12
        elif "AM" in hour_str and h == 12:
            h = 0
        rows[h] = {
            "netSales": float(m.group(2).replace(",","")),
            "guests": int(m.group(3)),
            "orders": int(m.group(4)),
            "laborHrs": float(m.group(5).replace(",","")),
            "laborDollars": float(m.group(6).replace(",","")),
            "laborPct": float(m.group(7).replace(",","")),
        }
    return rows


def load_teamworx_history(store: str) -> dict:
    """Returns {date: {sales: [18 hours], ideal, scheduled}} for the store."""
    # Try repo-local copy first (for GitHub Actions)
    candidates = [
        ROOT / "data" / f"teamworx_hourly_history_{store}.json",
        ROOT / "data" / "teamworx_mondays.json",
        CASE_STUDY / "teamworx_mondays.json",
    ]
    for f in candidates:
        if f.exists():
            data = json.loads(f.read_text())
            # Normalize: handle both flat {date: {...}} and nested {mondays: {date: {...}}}
            if "mondays" in data:
                return {date: {"sales": d["sales"], "ideal": d["ideal"], "scheduled": d["scheduled"]}
                        for date, d in data["mondays"].items()}
            return data
    return {}


def build_workbook(store: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    twx = load_teamworx_history(store)
    if not twx:
        print(f"No Teamworx history found for store {store}", file=sys.stderr)
        sys.exit(1)

    # Pull Brink for each date in twx
    brink_data = {}
    for date_str in sorted(twx.keys()):
        pdf = ROOT / "data" / "raw" / "parbrink" / store / date_str / "Hourly Sales And Labor.pdf"
        rows = parse_brink_pdf(pdf)
        if rows:
            brink_data[date_str] = rows

    print(f"Teamworx days: {len(twx)} | Brink days: {len(brink_data)} | Overlap: {len(set(twx) & set(brink_data))}")

    # Build Excel
    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill(start_color="DA291C", end_color="DA291C", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FILL = PatternFill(start_color="FFC72C", end_color="FFC72C", fill_type="solid")

    # Tab per overlapping date
    for date_str in sorted(set(twx) & set(brink_data)):
        ws = wb.create_sheet(date_str)
        ws.column_dimensions["A"].width = 16
        for i in range(2, 20):
            ws.column_dimensions[get_column_letter(i)].width = 9
        title = ws.cell(row=1, column=1, value=f"KY-{store} — {date_str} — Teamworx Forecast vs Brink Actual")
        title.font = Font(bold=True, size=14, color="DA291C")
        ws.merge_cells(f"A1:S1")

        # Header
        ws.cell(row=3, column=1, value="Metric").fill = HEADER_FILL
        ws.cell(row=3, column=1).font = HEADER_FONT
        for i, lbl in enumerate(LABELS):
            c = ws.cell(row=3, column=2+i, value=lbl)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT

        # Teamworx forecast
        ws.cell(row=4, column=1, value="TWX FORECAST $").font = Font(bold=True, color="0070C0")
        for i, v in enumerate(twx[date_str]["sales"]):
            ws.cell(row=4, column=2+i, value=v).number_format = "$#,##0"

        # Brink actual
        ws.cell(row=5, column=1, value="BRINK ACTUAL $").font = Font(bold=True, color="DA291C")
        ws.cell(row=5, column=1).fill = SECTION_FILL
        for i, h in enumerate(HOUR_24):
            v = brink_data[date_str].get(h, {}).get("netSales", 0)
            ws.cell(row=5, column=2+i, value=v).number_format = "$#,##0"

        # Variance $
        ws.cell(row=6, column=1, value="VARIANCE $ (Actual−Forecast)").font = Font(bold=True)
        for i, h in enumerate(HOUR_24):
            actual = brink_data[date_str].get(h, {}).get("netSales", 0)
            forecast = twx[date_str]["sales"][i]
            delta = actual - forecast
            cell = ws.cell(row=6, column=2+i, value=round(delta, 2))
            cell.number_format = "$#,##0"
            if delta < -50:
                cell.font = Font(color="DA291C", bold=True)
            elif delta > 50:
                cell.font = Font(color="0070C0", bold=True)

        # Variance %
        ws.cell(row=7, column=1, value="VARIANCE % (Δ/Forecast)").font = Font(bold=True)
        for i, h in enumerate(HOUR_24):
            actual = brink_data[date_str].get(h, {}).get("netSales", 0)
            forecast = twx[date_str]["sales"][i]
            if forecast > 0:
                pct = (actual - forecast) / forecast * 100
                cell = ws.cell(row=7, column=2+i, value=round(pct, 1))
                cell.number_format = "0.0%"
                ws.cell(row=7, column=2+i).value = pct / 100
                if pct < -15 or pct > 15:
                    cell.font = Font(color="DA291C", bold=True)

        # Brink labor hours + cost (the consulting story extension)
        ws.cell(row=9, column=1, value="BRINK Labor Hrs").font = Font(italic=True)
        ws.cell(row=10, column=1, value="BRINK Labor $").font = Font(italic=True)
        ws.cell(row=11, column=1, value="BRINK Labor %").font = Font(italic=True)
        for i, h in enumerate(HOUR_24):
            d = brink_data[date_str].get(h, {})
            ws.cell(row=9, column=2+i, value=d.get("laborHrs", 0))
            ws.cell(row=10, column=2+i, value=d.get("laborDollars", 0)).number_format = "$#,##0"
            ws.cell(row=11, column=2+i, value=d.get("laborPct", 0))

        ws.freeze_panes = "B4"

    # Summary tab
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 18
    for i in range(2, 22):
        ws.column_dimensions[get_column_letter(i)].width = 11

    title = ws.cell(row=1, column=1, value=f"KY-{store} — Teamworx Forecast Accuracy vs Brink Actual")
    title.font = Font(bold=True, size=14, color="DA291C")
    ws.merge_cells("A1:U1")

    # Summary table
    ws.cell(row=3, column=1, value="Date").fill = HEADER_FILL
    ws.cell(row=3, column=1).font = HEADER_FONT
    headers = ["Day", "TWX Total $", "Brink Total $", "Δ $", "Δ %", "Peak TWX hr", "Peak Brink hr", "Worst hr (Δ%)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=2+i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    overlap_dates = sorted(set(twx) & set(brink_data))
    row = 4
    for date_str in overlap_dates:
        twx_sales = twx[date_str]["sales"]
        brink_h = brink_data[date_str]
        twx_total = sum(twx_sales)
        brink_total = sum(brink_h.get(h, {}).get("netSales", 0) for h in HOUR_24)
        delta = brink_total - twx_total
        delta_pct = (delta / twx_total * 100) if twx_total else 0
        peak_twx = LABELS[twx_sales.index(max(twx_sales))] if max(twx_sales) > 0 else "—"
        brink_arr = [brink_h.get(h, {}).get("netSales", 0) for h in HOUR_24]
        peak_brink = LABELS[brink_arr.index(max(brink_arr))] if max(brink_arr) > 0 else "—"
        worst_idx = -1
        worst_pct = 0
        for i in range(len(HOUR_24)):
            if twx_sales[i] > 50:  # only care about real hours
                pct = (brink_arr[i] - twx_sales[i]) / twx_sales[i] * 100
                if abs(pct) > abs(worst_pct):
                    worst_pct = pct
                    worst_idx = i
        worst = f"{LABELS[worst_idx]} ({worst_pct:+.0f}%)" if worst_idx >= 0 else "—"

        from datetime import datetime as _dt
        dow = _dt.fromisoformat(date_str).strftime("%a")
        ws.cell(row=row, column=1, value=date_str)
        ws.cell(row=row, column=2, value=dow)
        ws.cell(row=row, column=3, value=twx_total).number_format = "$#,##0"
        ws.cell(row=row, column=4, value=brink_total).number_format = "$#,##0"
        ws.cell(row=row, column=5, value=delta).number_format = "$#,##0;[Red]-$#,##0"
        ws.cell(row=row, column=6, value=delta_pct/100).number_format = "0.0%"
        ws.cell(row=row, column=7, value=peak_twx)
        ws.cell(row=row, column=8, value=peak_brink)
        ws.cell(row=row, column=9, value=worst).font = Font(color="DA291C", bold=abs(worst_pct) > 20)
        row += 1

    out = CASE_STUDY / f"Teamworx_Forecast_vs_Brink_Actual_{store}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")
    print(f"Tabs: Summary + {len(overlap_dates)} day-detail tabs")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--store", default="2065")
    args = p.parse_args()
    build_workbook(args.store)
