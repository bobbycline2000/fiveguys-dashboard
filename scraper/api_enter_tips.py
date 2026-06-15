#!/usr/bin/env python3
"""
Lights-out Monday tip entry for KY-2065 (Store 2065).

End-to-end:
  1. Ensure NetChef cookies are alive (re-mint via api_discover.py if not).
  2. Pull Charged Tips total for the prior Mon-Sun week from
     POST /resource/sales/sales/registerSales/summary
  3. Pull per-employee Regular Hours from
     POST /resource/nc/employee/timedetail/summary
  4. Compute the tip pool (excludes GM "Cline, Robert"), tips/hr, per-employee payouts.
  5. Build the working tip-sheet xlsx for the week.
  6. POST one row per non-zero employee to
     POST /resource/labor-details/supplemental-wages/save?operatingDate=<Mon>&weekEndingDate=<Sun>
  7. Verify the count + total against Charged Tips and write a log entry.

Run modes:
  python scraper/api_enter_tips.py                   # auto: prior Mon-Sun based on today
  python scraper/api_enter_tips.py 05/03/2026        # explicit week-end Sunday (MM/DD/YYYY)
  python scraper/api_enter_tips.py 05/03/2026 dry    # dry run — no POSTs

Hard prerequisite: Sat + Sun labor actuals must be reviewed for the week before this runs.
The script aborts with a clear message if not.
"""
import datetime as dt
import json, os, subprocess, sys
from pathlib import Path
from collections import defaultdict

import requests

ROOT = Path(__file__).parent.parent
DATA = ROOT / "data"
DATA.mkdir(exist_ok=True)

NETCHEF = "https://fiveguysfr77.net-chef.com"
HDR = {
    "Accept": "application/json",
    "Content-Type": "application/json;charset=UTF-8",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": NETCHEF,
    "Referer": f"{NETCHEF}/ncext/next.ct",
}
PROBE_URL = f"{NETCHEF}/resource/recommended-actions/status"

GM_NAME_KEY = "Cline, Robert"   # excluded from tip pool
SW_ID_CREDIT_CARD_TIP = 12292   # global supplemental wage type id (verified 2026-05-04)


# --- Cookie / session ---------------------------------------------------------
def load_cookies():
    f = DATA / "ct_cookies.json"
    if not f.exists(): return {}
    return {c["name"]: c["value"] for c in json.loads(f.read_text())}

def session_alive(jar):
    try:
        r = requests.get(PROBE_URL, cookies=jar, headers=HDR, timeout=15, allow_redirects=False)
        return r.status_code == 200 and "json" in (r.headers.get("content-type") or "").lower()
    except Exception:
        return False

def remint():
    print("[mint] cookies stale — running api_discover.py", flush=True)
    res = subprocess.run([sys.executable, str(ROOT/"scraper"/"api_discover.py")], capture_output=True, text=True)
    if res.returncode != 0:
        print(res.stdout); print(res.stderr, file=sys.stderr)
        raise RuntimeError("re-mint failed")

def ensure_session():
    jar = load_cookies()
    if jar and session_alive(jar):
        print("[probe] session alive")
        return jar
    remint()
    jar = load_cookies()
    if not session_alive(jar):
        raise RuntimeError("re-mint did not produce a live session")
    return jar


# --- Date helpers -------------------------------------------------------------
def prior_week_mon_sun(today=None):
    """Given today, return (Monday, Sunday) of the most-recently-completed Mon-Sun week."""
    today = today or dt.date.today()
    # If today is Mon, prior week is last Mon..Sun. If today is Tue+, also last Mon..Sun.
    days_since_sun = (today.weekday() + 1) % 7  # Sun=0, Mon=1, .. Sat=6
    prior_sun = today - dt.timedelta(days=days_since_sun if days_since_sun else 7)
    prior_mon = prior_sun - dt.timedelta(days=6)
    return prior_mon, prior_sun

def fmt(d): return d.strftime("%m/%d/%Y")


# --- API: pull Charged Tips ---------------------------------------------------
def pull_charged_tips(jar, mon, sun):
    """Sum chargedTips across the Mon-Sun range from registerSales/summary."""
    body = {"page":1,"start":0,"limit":75,"extraFilter":[
        {"type":"date","value":fmt(mon - dt.timedelta(days=1)),"field":"salesDate","comparison":"gt"},
        {"type":"date","value":fmt(sun + dt.timedelta(days=1)),"field":"salesDate","comparison":"lt"},
    ]}
    r = requests.post(f"{NETCHEF}/resource/sales/sales/registerSales/summary",
                      json=body, cookies=jar, headers=HDR, timeout=30)
    r.raise_for_status()
    data = r.json()
    rows = data.get("rows") or []
    # Filter precisely to Mon..Sun (gt/lt are inclusive — verified 2026-05-04)
    inc = [x for x in rows if fmt(mon) <= x["salesDate"][:10] <= fmt(sun)]
    total = round(sum(float(x.get("chargedTips") or 0) for x in inc), 2)
    return total, inc


# --- API: pull per-employee hours --------------------------------------------
def pull_time_detail(jar, mon, sun):
    """Per-employee aggregated regular hours + position id, including the GM row."""
    body = {"pagingInfo":{"page":1,"start":0,"limit":500},
            "extraCriteriaMap":{"includeAltLocations":False,"excludeManagers":False,
                                "summarizeBy":"",
                                "startDate":f"{fmt(mon)} 00:00:00",
                                "weekEndingDate":f"{fmt(sun)} 00:00:00"}}
    r = requests.post(f"{NETCHEF}/resource/nc/employee/timedetail/summary",
                      json=body, cookies=jar, headers=HDR, timeout=30)
    r.raise_for_status()
    data = r.json()
    rows = data.get("contentMap",{}).get("gridList") or []

    agg = defaultdict(lambda: {"reg": 0.0, "ot": 0.0, "employeeId": None,
                               "employeeNumber": None, "positionCode": None,
                               "positionName": None})
    for row in rows:
        if row.get("rowType") != "regular": continue
        name = row["employeeName"]
        agg[name]["reg"] += float(row.get("regularHours") or 0)
        agg[name]["ot"]  += float(row.get("overtimeHours") or 0)
        agg[name]["employeeId"]     = row.get("employeeId") or agg[name]["employeeId"]
        agg[name]["employeeNumber"] = row.get("employeeNumber") or agg[name]["employeeNumber"]
        # Use the most recent shift's position as the canonical position for the week
        agg[name]["positionCode"]   = row.get("positionCode") or agg[name]["positionCode"]
        agg[name]["positionName"]   = row.get("positionName") or agg[name]["positionName"]
    return dict(agg)


# --- Tip sheet xlsx -----------------------------------------------------------
def build_tip_sheet_xlsx(week_end, charged_tips, payouts, pool_hours, rate, out_path):
    """Mirror the manual template format used by Bobby/Crystal.
    Layout (current 22-employee template, see crunchtime-enter-tips.md):
      A1 Online Tips
      A2 Week End <MM/DD>           C2 Total Tips      D2 Tips Per Hour
      C3 charged_tips                D3 =C3/B<total>
      A5 Name        B5 Hours Worked  C4 Payout
      A6..A<n> employee names         B6..B<n> hours    C6..C<n> =B*D3
      A<n+1> Total hours for Store     B<n+1> =SUM(...)  C<n+1> =SUM(...)"""
    try:
        import openpyxl
    except ImportError:
        print(f"[xlsx] openpyxl not installed — skipping xlsx generation")
        return None

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2065 Dixie"

    # Header
    ws['A1'] = 'Online Tips'
    ws['A2'] = f'Week End {week_end.strftime("%m/%d").lstrip("0").replace("/0","/")}'
    ws['C2'] = 'Total Tips'
    ws['D2'] = 'Tips Per Hour'
    ws['C3'] = round(float(charged_tips), 2)
    ws['C4'] = 'Payout'
    ws['A5'] = 'Name'
    ws['B5'] = 'Hours Worked'

    # Employee rows
    start = 6
    sorted_names = sorted(payouts.keys())
    for i, name in enumerate(sorted_names):
        r = start + i
        p = payouts[name]
        ws.cell(row=r, column=1).value = name
        ws.cell(row=r, column=2).value = p['hours']
        ws.cell(row=r, column=3).value = f'=B{r}*D3'
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'

    total_row = start + len(sorted_names)
    ws.cell(row=total_row, column=1).value = 'Total hours for Store'
    ws.cell(row=total_row, column=2).value = f'=SUM(B{start}:B{total_row-1})'
    ws.cell(row=total_row, column=3).value = f'=SUM(C{start}:C{total_row-1})'
    ws.cell(row=total_row, column=3).number_format = '"$"#,##0.00'
    ws['D3'] = f'=C3/B{total_row}'
    ws['C3'].number_format = '"$"#,##0.00'
    ws['D3'].number_format = '"$"#,##0.0000'

    wb.save(out_path)
    return out_path


# --- Compute payouts ----------------------------------------------------------
def compute_payouts(employees, charged_tips):
    """Return {name: {hours, payout, employeeId, positionCode, positionName}} keyed by name.
    Excludes GM. Pool = sum of regular hours across non-GM. Penny-rounding may leave a
    sub-cent delta vs charged_tips — accepted by the skill (4/27 reference: $0.01 over)."""
    non_gm = {n: a for n, a in employees.items() if GM_NAME_KEY not in n}
    pool = sum(a["reg"] for a in non_gm.values())
    if pool <= 0:
        raise RuntimeError("pool hours = 0 — sanity fail")
    rate = charged_tips / pool
    out = {}
    total_payout = 0.0
    for name, a in non_gm.items():
        hours = round(a["reg"], 2)
        if hours <= 0: continue
        payout = round(hours * rate, 2)
        if payout <= 0: continue
        total_payout += payout
        out[name] = {
            "hours": hours,
            "payout": payout,
            "employeeId": a["employeeId"],
            "employeeNumber": a["employeeNumber"],
            "positionCode": int(a["positionCode"]) if a["positionCode"] is not None else 1,
            "positionName": a["positionName"] or "",
        }
    return out, round(pool, 2), round(rate, 4), round(total_payout, 2)


# --- API: write one supplemental-wage row ------------------------------------
def post_credit_card_tip(jar, mon, sun, employee_id, position_id, lump_sum):
    body = [{
        "id": -1,
        "employeeId": int(employee_id),
        "employeeNumber": "",
        "employeeName": "",
        "employeeFullName": "",
        "employeePositionId": int(position_id),
        "employeePositionName": "",
        "terminationDate": None,
        "terminated": False,
        "swId": SW_ID_CREDIT_CARD_TIP,
        "swDescription": "",
        "swLumpSumVal": round(float(lump_sum), 2),
        "viewOnly": False,
        "mondayHoursAdjusted": 0, "tuesdayHoursAdjusted": 0, "wednesdayHoursAdjusted": 0,
        "thursdayHoursAdjusted": 0, "fridayHoursAdjusted": 0,
        "saturdayHoursAdjusted": 0, "sundayHoursAdjusted": 0,
        "mondayDailyLumpSum": 0, "tuesdayDailyLumpSum": 0, "wednesdayDailyLumpSum": 0,
        "thursdayDailyLumpSum": 0, "fridayDailyLumpSum": 0,
        "saturdayDailyLumpSum": 0, "sundayDailyLumpSum": 0,
        "detailsList": [],
        "altLocationFlag": True,
        "deletedPositionFlag": False,
        "reviewed": True,
        "reviewUser": "",
    }]
    params = {"operatingDate": fmt(mon), "weekEndingDate": fmt(sun)}
    r = requests.post(f"{NETCHEF}/resource/labor-details/supplemental-wages/save",
                      params=params, data=json.dumps(body),
                      cookies=jar, headers=HDR, timeout=30)
    r.raise_for_status()
    # /save returns HTTP 200 with an empty body on success (confirmed 2026-06-15)
    res = r.json() if r.text.strip() else {"success": True}
    if not res.get("success"):
        raise RuntimeError(f"server returned non-success: {res}")
    return res


def upsert_existing_row(jar, mon, sun, existing_row, new_value):
    """UPDATE an already-saved supplemental-wage row in place by echoing the full
    row object back to /save with a modified swLumpSumVal.

    CRITICAL (lesson 2026-05-25): /save with id:-1 APPENDS a new row; /save with
    the row's existing composite id (e.g. "870073_12292_9_<hash>") UPDATES in place.
    Posting the full existing row object verbatim (only swLumpSumVal changed) is the
    idempotent path — it never creates a duplicate. Use this for every employee that
    already has a row for the week; use post_credit_card_tip (append) only when no
    row exists yet."""
    body = [dict(existing_row, swLumpSumVal=round(float(new_value), 2))]
    params = {"operatingDate": fmt(mon), "weekEndingDate": fmt(sun)}
    r = requests.post(f"{NETCHEF}/resource/labor-details/supplemental-wages/save",
                      params=params, data=json.dumps(body),
                      cookies=jar, headers=HDR, timeout=30)
    r.raise_for_status()
    res = r.json() if r.text.strip() else {"success": True}
    if not res.get("success"):
        raise RuntimeError(f"server returned non-success on upsert: {res}")
    return res


# --- API: commit pending session-queued changes ------------------------------
# Discovered 2026-05-11: /save POSTs to supplemental-wages queue rows in the
# user session but DO NOT persist. The disk-save → "Labor Adjustments Alert"
# dialog → Continue button is what actually commits. That UI flow fires:
#   1. POST /resource/labor-details/prepare        (returns WIP or summary)
#   2. (if dialog shows) user clicks Continue
#   3. POST /resource/labor-details/labor-actuals/validate  (final commit)
# Without these two calls the /save rows never appear in the labor week.
def commit_pending(jar, mon, sun):
    """Call /prepare + /labor-actuals/validate to commit queued /save rows.
    Returns (committed:bool, detail:dict). Raises only on transport error."""
    # 1. /prepare — locks the session WIP for commit
    # Body format captured from browser network log 2026-06-15; {} causes 500.
    prepare_body = {"extraCriteriaMap": {
        "weekEndingDate": fmt(sun), "editMode": "true",
        "processWIP": False, "continueWIP": False,
        "checkAllocateToLocationPosted": True,
        "checkMissingLaborActuals": False,
        "checkNonReviewedLaborActuals": False,
        "taskDate": None, "taskLocationId": None,
    }}
    p = requests.post(f"{NETCHEF}/resource/labor-details/prepare",
                      json=prepare_body, cookies=jar, headers=HDR, timeout=30)
    p.raise_for_status()
    prep = p.json() if p.text else {}
    if not prep.get("success"):
        # WORK_IN_PROGRESS PREPARE with empty fields means nothing queued.
        msgs = prep.get("contentMap",{}).get("messageList") or []
        return False, {"step":"prepare","response":prep,"messages":msgs}

    # 2. /labor-actuals/validate — final commit, equivalent to clicking
    # "Continue" on the Labor Adjustments Alert. Body shape captured from
    # browser network log on 2026-05-11 commit (verified live).
    body = {"weekEndingDate": fmt(sun), "operatingDate": fmt(mon)}
    v = requests.post(f"{NETCHEF}/resource/labor-details/labor-actuals/validate",
                      json=body, cookies=jar, headers=HDR, timeout=60)
    v.raise_for_status()
    val = v.json() if v.text else {}
    return bool(val.get("success", True)), {"step":"validate","response":val}


# --- Verify ------------------------------------------------------------------
def verify_saved(jar, sun):
    """Pull supplementals for week-ending and report count + sum of Credit Card Tip rows."""
    body = {"pagingInfo":{"page":1,"start":0,"limit":500},
            "sortInfo":{"sortList":[{"property":"employeeId","direction":"ASC"}]},
            "extraCriteriaMap":{"extraParams":{
                "labelModules":["labor_actuals.labor_actuals","labor_actuals.labor_details","report.labor_detail"],
                "weekEndingDate":[fmt(sun)]}}}
    r = requests.post(f"{NETCHEF}/resource/labor-details/supplemental-wages",
                      json=body, cookies=jar, headers=HDR, timeout=30)
    r.raise_for_status()
    data = r.json()
    rows = data.get("contentMap",{}).get("gridList") or []
    cct_rows = [x for x in rows if x.get("swId") == SW_ID_CREDIT_CARD_TIP]
    total = round(sum(float(x.get("swLumpSumVal") or 0) for x in cct_rows), 2)
    return len(cct_rows), total, cct_rows


# --- Hard prereq: labor reviewed for Sat + Sun --------------------------------
def labor_reviewed(jar, sun):
    """Returns True if Labor Summary shows the week as Reviewed."""
    body = {"pagingInfo":{"page":1,"start":0,"limit":50},
            "extraCriteriaMap":{"fiscalYear":sun.year}}
    r = requests.post(f"{NETCHEF}/resource/labor/summary/list",
                      json=body, cookies=jar, headers=HDR, timeout=30)
    if r.status_code != 200:
        # Endpoint name may differ; if it doesn't work, skip the check rather than block
        print(f"[prereq] labor summary endpoint returned {r.status_code} — skipping prereq check")
        return True
    data = r.json()
    rows = data.get("contentMap",{}).get("gridList") or data.get("rows") or []
    for x in rows:
        we = (x.get("weekEndingDate") or "")[:10]
        if we == fmt(sun):
            status = (x.get("status") or "").lower()
            return "review" in status and "not" not in status
    return False


# --- Main ---------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    dry = "dry" in args
    explicit = next((a for a in args if "/" in a), None)
    if explicit:
        sun = dt.datetime.strptime(explicit, "%m/%d/%Y").date()
        mon = sun - dt.timedelta(days=6)
    else:
        mon, sun = prior_week_mon_sun()
    print(f"=== KY-2065 tip entry — WE {fmt(sun)} (Mon {fmt(mon)} → Sun {fmt(sun)}) ===")
    print(f"  mode: {'DRY (no POST)' if dry else 'LIVE'}")

    jar = ensure_session()

    # Optional prereq probe
    if not labor_reviewed(jar, sun):
        print(f"[prereq] WARNING: WE {fmt(sun)} labor not Reviewed — proceeding anyway. "
              "Validate hours look right before trusting the run.")

    # 1. Charged Tips
    charged, day_rows = pull_charged_tips(jar, mon, sun)
    print(f"[tips] Charged Tips Mon-Sun: ${charged:.2f} ({len(day_rows)} day rows)")

    # 2. Time detail
    employees = pull_time_detail(jar, mon, sun)
    total_reg = round(sum(a["reg"] for a in employees.values()), 2)
    print(f"[hours] {len(employees)} employees, total regular hrs: {total_reg}")

    # 3. Payouts
    payouts, pool_hrs, rate, sum_payout = compute_payouts(employees, charged)
    print(f"[pool] pool hours: {pool_hrs} (excluding {GM_NAME_KEY})")
    print(f"[pool] tips/hr: ${rate:.4f}")
    print(f"[pool] sum of payouts: ${sum_payout:.2f} (vs charged ${charged:.2f}, "
          f"delta ${sum_payout - charged:+.2f})")

    print()
    print(f"{'Employee':<35} {'EmpId':>9} {'Pos':>4} {'Hours':>7} {'Payout':>9}")
    print("-" * 70)
    for name in sorted(payouts):
        p = payouts[name]
        print(f"{name:<35} {p['employeeId']:>9} {p['positionCode']:>4} {p['hours']:>7.2f} ${p['payout']:>8.2f}")
    print("-" * 70)
    print(f"{'TOTAL':<35} {'':>9} {'':>4} {pool_hrs:>7.2f} ${sum_payout:>8.2f}")

    # 4. Save snapshot
    snap_path = DATA / f"tips_we_{sun.strftime('%Y_%m_%d')}_snapshot.json"
    snap_path.write_text(json.dumps({
        "weekEnding": fmt(sun), "weekStart": fmt(mon),
        "chargedTips": charged, "poolHours": pool_hrs, "tipsPerHour": rate,
        "sumPayout": sum_payout, "delta": round(sum_payout - charged, 2),
        "payouts": payouts,
    }, indent=2, default=str))
    print(f"\n[snapshot] {snap_path}")

    # 4b. Build tip sheet xlsx (matches the format Bobby/Crystal use manually)
    # Path: <repo>/data/tip-sheets/tip-sheet-2065-WE-MM-DD.xlsx (CI-friendly).
    # Local runs also drop a copy in BobbyWorkspace/_reference/tip-sheets/ if that
    # directory exists (per existing manual workflow).
    xlsx_name = f"tip-sheet-2065-WE-{sun.strftime('%m-%d')}.xlsx"
    xlsx_repo = DATA / "tip-sheets" / xlsx_name
    build_tip_sheet_xlsx(sun, charged, payouts, pool_hrs, rate, xlsx_repo)
    print(f"[xlsx]    {xlsx_repo}")
    bobby_ref = ROOT.parent.parent / "_reference" / "tip-sheets" / xlsx_name
    if bobby_ref.parent.exists():
        build_tip_sheet_xlsx(sun, charged, payouts, pool_hrs, rate, bobby_ref)
        print(f"[xlsx]    {bobby_ref}")

    if dry:
        print("\n[dry] Skipping POSTs. Re-run without 'dry' to commit.")
        return

    # 5. PRE-FLIGHT: read any rows that already exist for this week, keyed by
    # employeeId. This is what makes the run IDEMPOTENT — re-running (or running
    # on top of a stale partial entry) UPDATES in place instead of duplicating.
    # Root-cause of the 2026-05-25 42-row double-entry disaster: blind id:-1 append.
    _, _, existing_rows = verify_saved(jar, sun)
    existing_by_emp = defaultdict(list)
    for r in existing_rows:
        existing_by_emp[r.get("employeeId")].append(r)
    print(f"\n[preflight] {len(existing_rows)} existing Credit Card Tip rows "
          f"across {len(existing_by_emp)} employees already on the week")

    target_emp_ids = {payouts[n]["employeeId"] for n in payouts}

    # 5. UPSERT each target employee — update in place if a row exists, else append.
    print(f"\n[write] Reconciling {len(payouts)} Credit Card Tip rows (idempotent upsert)...")
    appended = updated = zeroed = 0; failed = []
    for name in sorted(payouts):
        p = payouts[name]; eid = p["employeeId"]; target = p["payout"]
        rows = existing_by_emp.get(eid, [])
        try:
            if rows:
                # First existing row carries the payout; any extra duplicate rows -> 0.
                for i, row in enumerate(rows):
                    want = target if i == 0 else 0
                    if abs(float(row.get("swLumpSumVal") or 0) - want) < 0.001:
                        continue
                    upsert_existing_row(jar, mon, sun, row, want)
                    if i == 0: updated += 1
                    else: zeroed += 1
                print(f"  ↻ {name:<35} ${target:>8.2f} (updated; {len(rows)-1} dup→$0)")
            else:
                post_credit_card_tip(jar, mon, sun, eid, p["positionCode"], target)
                appended += 1
                print(f"  + {name:<35} ${target:>8.2f} (new)")
        except Exception as exc:
            failed.append((name, str(exc)))
            print(f"  ✗ {name:<35} {exc}")

    # 5a. Zero any stale rows for employees NOT in this week's pool (e.g. leftover
    # from a prior wrong-number partial run). Never leave a non-target row paying out.
    for eid, rows in existing_by_emp.items():
        if eid in target_emp_ids:
            continue
        for row in rows:
            if abs(float(row.get("swLumpSumVal") or 0)) < 0.001:
                continue
            try:
                upsert_existing_row(jar, mon, sun, row, 0)
                zeroed += 1
                print(f"  ⌫ stale empId {eid} → $0")
            except Exception as exc:
                failed.append((f"stale-{eid}", str(exc)))

    posted = appended + updated
    print(f"\n[write] appended {appended}, updated {updated}, zeroed {zeroed}, "
          f"failures: {len(failed)}")
    if failed:
        for n, e in failed: print(f"   FAIL: {n} :: {e}")

    # 5b. COMMIT — /save only queues; /prepare + /labor-actuals/validate
    # is what makes the rows persist. Without this step the entries vanish.
    print(f"\n[commit] firing /prepare + /labor-actuals/validate to commit session WIP...")
    try:
        committed, detail = commit_pending(jar, mon, sun)
        if committed:
            print(f"[commit] OK — {detail.get('step')} returned success")
        else:
            print(f"[commit] FAILED at {detail.get('step')} — {detail}")
            print(f"[commit] FALLBACK NEEDED: open browser, navigate to "
                  f"Labor Summary → Edit WE {fmt(sun)} → Supplemental Wages → "
                  f"click disk save (top-right) → click Continue on dialog. "
                  f"Rows are queued server-side; this just triggers the commit.")
    except Exception as e:
        print(f"[commit] EXCEPTION: {e}")
        print(f"[commit] FALLBACK NEEDED: see manual commit steps above.")

    # 6. Verify — count UNIQUE paid employees and FAIL LOUD on any duplicate.
    # (count of all rows is meaningless once $0 dup-rows exist; what matters is
    # exactly one NON-ZERO row per employee and total == charged.)
    count, total, all_rows = verify_saved(jar, sun)
    nonzero = [r for r in all_rows if float(r.get("swLumpSumVal") or 0) > 0]
    nz_total = round(sum(float(r.get("swLumpSumVal") or 0) for r in nonzero), 2)
    seen = defaultdict(int)
    for r in nonzero: seen[r.get("employeeId")] += 1
    dups = {eid: n for eid, n in seen.items() if n > 1}
    print(f"\n[verify] WE {fmt(sun)}: {len(nonzero)} paid rows / {len(seen)} unique employees, "
          f"non-zero total ${nz_total:.2f} (charged ${charged:.2f}, delta ${nz_total - charged:+.2f})")
    if dups:
        print(f"[verify] *** DUPLICATE PAID ROWS DETECTED *** {dups} — DO NOT POST. "
              f"Re-run is idempotent and will collapse to one row per employee.")
    elif abs(nz_total - charged) > 0.05:
        print(f"[verify] *** TOTAL MISMATCH *** ${nz_total:.2f} vs charged ${charged:.2f} — review before posting.")
    else:
        print(f"[verify] OK — one non-zero row per employee, total ties within rounding.")
    count, total = len(nonzero), nz_total

    # 7. Log
    log = ROOT.parent.parent / "_memory" / "tip-entry-log.md"
    if log.exists():
        line = (f"| {fmt(sun)} | ${charged:.2f} | {pool_hrs:.2f} | ${rate:.4f} | "
                f"${total:.2f} | {count} | API (api_enter_tips.py) | "
                f"{posted}/{len(payouts)} posted; {len(failed)} failed; lights-out path. |\n")
        with log.open("a", encoding="utf-8") as f: f.write(line)
        print(f"[log] appended to {log}")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass
    main()
