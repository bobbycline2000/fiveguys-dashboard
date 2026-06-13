"""
Five Guys 2065 Secret Shopper Tracker — v3
Matches original training-tracker format exactly:
  Row 1 : Title "SECRET SHOPPER SCORES — [MONTH YEAR]"
  Row 2 : WEEK 1 / WEEK 2 / … headers + MTD SCORE / MTD DOLLARS
  Row 3 : Date
  Row 4 : Meal Period
  Row 5 : TOTAL SCORE (integer %)
  Rows 6-10 : Sub-score labels (blank — shops.json lacks those fields)
  Row 11: INDIVIDUAL PERFORMANCE divider + MTD column headers
  Row 12+: Manager rows (col A="Manager"), then Crew rows (col A="CREW")
            col B = display name, shop cols = score if on shift else blank,
            MTD SCORE = AVERAGE formula, MTD DOLLARS = Python-computed

Bug fixes vs v2:
  * Year filter: ONLY 2026 (date >= 2026-01-01). Eliminates 2025-Jun bleed-in.
  * Employee scores: integer score value or blank — not "Y" / "N".
  * MTD DOLLARS: $230 / len(on_shift) per 100%-shop worked, summed.
  * Roster: current active list from scripts/build_employee_directory.py.
  * Name normalization: "Zack" → "Zach" etc. so participation.json matches roster keys.
"""

import json
import os
from collections import defaultdict
from datetime import date as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── Colours ────────────────────────────────────────────────────────────────
RED         = "C8102E"
WHITE       = "FFFFFF"
OFF_WHITE   = "F9F9F9"
LIGHT_GREY  = "F2F2F2"
MID_GREY    = "D0D0D0"
DARK        = "1A1A1A"
DARK_GREY   = "595959"
GREEN       = "2E7D32"
LIGHT_GREEN = "E8F5E9"
GOLD        = "D4A017"
LIGHT_GOLD  = "FFFBF0"
RED_LIGHT   = "FDECEA"
NAVY        = "1A2A4A"
MGR_BG      = "FCE7C7"

# ── Paths ──────────────────────────────────────────────────────────────────
HERE      = Path(__file__).resolve().parent
DASH_ROOT = HERE.parent
DATA_ROOT = DASH_ROOT / "data" / "raw" / "marketforce"
STORE_ID  = os.environ.get("STORE_ID", "2065")
OUT       = DASH_ROOT / "Shop_Tracker_2065_2026.xlsx"


def _find_latest_shops_json() -> Path:
    store_dir = DATA_ROOT / STORE_ID
    candidates = sorted(
        (x for x in store_dir.iterdir() if x.is_dir()), reverse=True
    )
    for d in candidates:
        p = d / "shops.json"
        if p.exists():
            return p
    raise FileNotFoundError(f"No shops.json found under {store_dir}")


SHOPS_JSON = _find_latest_shops_json()
PART_JSON  = DATA_ROOT / STORE_ID / "participation.json"

# ── Employee roster ────────────────────────────────────────────────────────
# (participation_key, display_name, section)
# participation_key must match first-name keys in participation.json (after normalization)
MANAGERS = [
    ("Vicki",  "Vicki Lucey",    "Manager"),
    ("Kasey",  "Kasey Wilson",   "Manager"),
    ("Nathan", "Nathan Roberts", "Manager"),
]

CREW = [
    ("Alen",      "Ailen Gonzalez Cardero",       "CREW"),
    ("Anthony",   "Anthony Pirtle",                "CREW"),
    ("Ash",       "Ash",                           "CREW"),
    ("Autumn",    "Autumn Hearn",                  "CREW"),
    ("Bri",       "Brianna Gatewood",              "CREW"),
    ("Dakayla",   "DaKayla Dorsey",                "CREW"),
    ("Damon",     "Damon",                         "CREW"),
    ("Divan",     "Divan",                         "CREW"),
    ("Francisco", "Francisco Llorente Mejias",     "CREW"),
    ("Grace",     "Grace",                         "CREW"),
    ("Jada",      "Jada Cox",                      "CREW"),
    ("Jeremiah",  "Jeremiah Matthews",             "CREW"),
    ("Kable",     "Kable",                         "CREW"),
    ("Kaisha",    "Kaisha Brewer",                 "CREW"),
    ("Kayla",     "Kayla Valenzuela",              "CREW"),
    ("Kenzie",    "Kenzie",                        "CREW"),
    ("Lidy",      "Lidy Henry",                    "CREW"),
    ("Maylin",    "Maylin Hernandez Rodriguez",    "CREW"),
    ("Mike",      "Mike",                          "CREW"),
    ("Nyatiek",   "Nyatiek Keah",                  "CREW"),
    ("Richard",   "Richard Gibbs",                 "CREW"),
    ("Ryan",      "Ryan",                          "CREW"),
    ("Samuel",    "Samuel Galban Rocha",           "CREW"),
    ("Zach",      "Zack Whitten",                  "CREW"),
]

# ── Name normalisation: participation.json first-name → roster key ─────────
# Handles alternate spellings that appear in participation.json
PART_KEY_NORM: dict[str, str] = {
    "Zack":    "Zach",
    "DaKayla": "Dakayla",
    "Brianna": "Bri",
}

MONTHS = {
    1: "January", 2: "February", 3: "March",    4: "April",
    5: "May",     6: "June",     7: "July",      8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

SUB_ROWS = [
    "Overall Experience",
    "Enthusiasm",
    "Likely to Return",
    "Likely to Recommend",
    "Exceptional Act?",
]


# ── Helpers ────────────────────────────────────────────────────────────────

def _norm(name: str) -> str:
    return PART_KEY_NORM.get(name, name)


def _b(colour: str = MID_GREY) -> Border:
    s = Side(border_style="thin", color=colour)
    return Border(left=s, right=s, top=s, bottom=s)


def _f(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour, end_color=colour)


def _set(cell, value=None, bold=False, size=10, fg=DARK, bg=None,
         h="center", v="center", wrap=False, italic=False,
         border=True, num_fmt=None) -> None:
    if value is not None:
        cell.value = value
    cell.font = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill = _f(bg)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        cell.border = _b()
    if num_fmt:
        cell.number_format = num_fmt


def _score_bg(s) -> str:
    if s is None:  return OFF_WHITE
    if s >= 100:   return LIGHT_GREEN
    if s >= 90:    return "EFF8E8"
    if s >= 80:    return LIGHT_GOLD
    return RED_LIGHT


def _score_fg(s) -> str:
    if s is None:  return DARK
    if s >= 100:   return GREEN
    if s >= 90:    return "4A7A0F"
    if s >= 80:    return GOLD
    return RED


def _blank_cell(ws, r: int, col_letter: str, bg: str = OFF_WHITE) -> None:
    """Set styling on a cell without writing any value (leaves it truly empty)."""
    c = ws.cell(r, column_index_from_string(col_letter))
    c.fill  = _f(bg)
    c.border = _b()
    c.alignment = Alignment(horizontal="center", vertical="center")


def compute_mtd_dollars(month_shops: list, raw_by_shop: dict,
                        roster_key: str) -> float:
    """$230 / n_on_shift for every 100 %-shop the employee worked this month."""
    total = 0.0
    for shop in month_shops:
        if shop["score"] != 100.0:
            continue
        jid = shop["job_id"]
        raw_names = raw_by_shop.get(jid, [])
        if not raw_names:
            continue
        norm_names = {_norm(n) for n in raw_names}
        if roster_key in norm_names:
            total += 230.0 / len(raw_names)   # divisor = actual on-shift count
    return round(total, 2)


# ── Sheet builder ──────────────────────────────────────────────────────────

def build_sheet(wb: Workbook, month_name: str, shops: list,
                raw_by_shop: dict) -> None:
    ws = wb.create_sheet(month_name)
    ws.sheet_view.showGridLines = False

    n = len(shops)
    # Columns: A=1, B=2, C...(2+n)=shop cols, (3+n)=MTD SCORE, (4+n)=MTD DOLLARS
    shop_cols      = [get_column_letter(3 + i) for i in range(n)]
    mtd_score_col  = get_column_letter(3 + n)
    mtd_dollar_col = get_column_letter(4 + n)
    last_col       = mtd_dollar_col

    # Column widths
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 24
    for col in shop_cols:
        ws.column_dimensions[col].width = 8
    ws.column_dimensions[mtd_score_col].width  = 10
    ws.column_dimensions[mtd_dollar_col].width = 13

    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    r = 1

    # ── Row 1: Title ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:{last_col}{r}")
    _set(ws.cell(r, 1),
         f"SECRET SHOPPER SCORES — {month_name.upper()} 2026",
         bold=True, size=14, fg=WHITE, bg=RED, border=False, h="center")
    ws.row_dimensions[r].height = 28
    r += 1   # → 2

    # ── Row 2: WEEK / MTD column headers ───────────────────────────────────
    _set(ws.cell(r, 1), "",          fg=WHITE, bg=NAVY)
    _set(ws.cell(r, 2), "",          fg=WHITE, bg=NAVY)
    for i, col in enumerate(shop_cols):
        _set(ws.cell(r, column_index_from_string(col)),
             f"WEEK {i + 1}", bold=True, size=9, fg=WHITE, bg=NAVY)
    _set(ws.cell(r, column_index_from_string(mtd_score_col)),
         "MTD SCORE",  bold=True, size=9, fg=WHITE, bg=NAVY)
    _set(ws.cell(r, column_index_from_string(mtd_dollar_col)),
         "MTD DOLLARS", bold=True, size=9, fg=WHITE, bg=NAVY)
    ws.row_dimensions[r].height = 18
    r += 1   # → 3

    # ── Row 3: Date ─────────────────────────────────────────────────────────
    _set(ws.cell(r, 1), "",     bg=LIGHT_GREY)
    _set(ws.cell(r, 2), "Date", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="left")
    for i, col in enumerate(shop_cols):
        d_str = dt.fromisoformat(shops[i]["date"]).strftime("%m/%d")
        _set(ws.cell(r, column_index_from_string(col)),
             d_str, bold=True, size=9, fg=DARK, bg=LIGHT_GREY)
    _set(ws.cell(r, column_index_from_string(mtd_score_col)),  "", bg=LIGHT_GREY)
    _set(ws.cell(r, column_index_from_string(mtd_dollar_col)), "", bg=LIGHT_GREY)
    ws.row_dimensions[r].height = 15
    r += 1   # → 4

    # ── Row 4: Meal Period ──────────────────────────────────────────────────
    _set(ws.cell(r, 1), "",           bg=LIGHT_GREY)
    _set(ws.cell(r, 2), "Meal Period", bold=True, size=9, fg=DARK, bg=LIGHT_GREY, h="left")
    for i, col in enumerate(shop_cols):
        mp = shops[i]["meal_period"]
        is_lunch = "lunch" in mp.lower()
        label = "Lunch" if is_lunch else "Dinner"
        bg_  = LIGHT_GOLD if is_lunch else "E8EEF8"
        fg_  = GOLD       if is_lunch else NAVY
        _set(ws.cell(r, column_index_from_string(col)),
             label, bold=True, size=8, fg=fg_, bg=bg_)
    _set(ws.cell(r, column_index_from_string(mtd_score_col)),  "", bg=LIGHT_GREY)
    _set(ws.cell(r, column_index_from_string(mtd_dollar_col)), "", bg=LIGHT_GREY)
    ws.row_dimensions[r].height = 15
    r += 1   # → 5

    # ── Row 5: TOTAL SCORE ──────────────────────────────────────────────────
    score_row = r
    _set(ws.cell(r, 1), "",            bg=RED_LIGHT)
    _set(ws.cell(r, 2), "TOTAL SCORE", bold=True, size=10, fg=RED, bg=RED_LIGHT, h="left")
    for i, col in enumerate(shop_cols):
        pct = shops[i]["score"]
        val = int(pct) if pct == int(pct) else pct
        _set(ws.cell(r, column_index_from_string(col)),
             val, bold=True, size=11, fg=_score_fg(pct), bg=_score_bg(pct))
    _set(ws.cell(r, column_index_from_string(mtd_score_col)),  "", bg=RED_LIGHT)
    _set(ws.cell(r, column_index_from_string(mtd_dollar_col)), "", bg=RED_LIGHT)
    ws.row_dimensions[r].height = 20
    r += 1   # → 6

    # ── Rows 6-10: Sub-score categories (labels only — no data in shops.json) ─
    for label in SUB_ROWS:
        _set(ws.cell(r, 1), "",    bg=OFF_WHITE)
        _set(ws.cell(r, 2), label, size=9, fg=DARK_GREY, bg=OFF_WHITE, h="left")
        for col in shop_cols:
            _blank_cell(ws, r, col, bg=OFF_WHITE)
        _blank_cell(ws, r, mtd_score_col,  bg=OFF_WHITE)
        _blank_cell(ws, r, mtd_dollar_col, bg=OFF_WHITE)
        ws.row_dimensions[r].height = 14
        r += 1   # → 11 after 5 iterations

    # ── Row 11: INDIVIDUAL PERFORMANCE divider ─────────────────────────────
    ws.merge_cells(f"A{r}:B{r}")
    _set(ws.cell(r, 1),
         "INDIVIDUAL PERFORMANCE", bold=True, size=10, fg=WHITE, bg=NAVY, h="left")
    for col in shop_cols:
        _set(ws.cell(r, column_index_from_string(col)), "", fg=WHITE, bg=NAVY)
    _set(ws.cell(r, column_index_from_string(mtd_score_col)),
         "MTD SCORE",  bold=True, size=9, fg=WHITE, bg=NAVY)
    _set(ws.cell(r, column_index_from_string(mtd_dollar_col)),
         "MTD DOLLARS", bold=True, size=9, fg=WHITE, bg=NAVY)
    ws.row_dimensions[r].height = 20
    r += 1   # → 12

    # ── Pre-build normalised on-shift lookup for this month's shops ─────────
    norm_on_shift: dict[str, set] = {}
    for shop in shops:
        jid = shop["job_id"]
        raw = raw_by_shop.get(jid, [])
        norm_on_shift[jid] = {_norm(n) for n in raw}

    # First/last shop column letters for AVERAGE formula
    if shop_cols:
        first_shop_col = shop_cols[0]
        last_shop_col  = shop_cols[-1]

    # ── Employee rows (Managers, then Crew) ─────────────────────────────────
    for section_label, emp_list in (("Manager", MANAGERS), ("CREW", CREW)):
        is_mgr = section_label == "Manager"
        row_bg = MGR_BG if is_mgr else WHITE

        for roster_key, display_name, _sec in emp_list:
            _set(ws.cell(r, 1), section_label,
                 size=8, fg=DARK_GREY, bg=row_bg, bold=is_mgr)
            _set(ws.cell(r, 2), display_name,
                 size=10, fg=DARK, bg=row_bg, h="left", bold=is_mgr)

            for i, col in enumerate(shop_cols):
                shop = shops[i]
                jid  = shop["job_id"]
                if roster_key in norm_on_shift.get(jid, set()):
                    pct = shop["score"]
                    val = int(pct) if pct == int(pct) else pct
                    _set(ws.cell(r, column_index_from_string(col)),
                         val, bold=True, size=10,
                         fg=_score_fg(pct), bg=_score_bg(pct))
                else:
                    _blank_cell(ws, r, col, bg=row_bg)

            # MTD SCORE: Excel AVERAGE() skips empty cells natively
            if shop_cols:
                formula = (f"=IFERROR(AVERAGE("
                           f"{first_shop_col}{r}:{last_shop_col}{r}),\"\")")
                mtd_cell = ws.cell(r, column_index_from_string(mtd_score_col))
                mtd_cell.value         = formula
                mtd_cell.font          = Font(name="Arial", bold=True, size=9, color=DARK)
                mtd_cell.fill          = _f(LIGHT_GREY)
                mtd_cell.alignment     = Alignment(horizontal="center", vertical="center")
                mtd_cell.border        = _b()
                mtd_cell.number_format = "0.0"
            else:
                _blank_cell(ws, r, mtd_score_col)

            # MTD DOLLARS: Python-computed (divisor varies per shop)
            dollars = compute_mtd_dollars(shops, raw_by_shop, roster_key)
            dc = ws.cell(r, column_index_from_string(mtd_dollar_col))
            if dollars > 0:
                dc.value          = dollars
                dc.font           = Font(name="Arial", bold=True, size=10, color=GREEN)
                dc.fill           = _f(LIGHT_GREEN)
                dc.alignment      = Alignment(horizontal="center", vertical="center")
                dc.border         = _b()
                dc.number_format  = '"$"#,##0.00'
            else:
                _blank_cell(ws, r, mtd_dollar_col, bg=row_bg)

            ws.row_dimensions[r].height = 17
            r += 1

    # ── Footer ─────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:{last_col}{r}")
    _set(ws.cell(r, 1),
         "Score = shop score if on shift during meal window; blank if not. "
         "MTD SCORE = average of scores worked. "
         "MTD DOLLARS = $230 ÷ on-shift count per 100% shop.",
         size=7, fg=DARK_GREY, bg=LIGHT_GREY, italic=True, h="left", border=False)
    ws.row_dimensions[r].height = 12


# ── Main ───────────────────────────────────────────────────────────────────

def build() -> None:
    shops_data = json.loads(SHOPS_JSON.read_text(encoding="utf-8"))
    all_shops  = sorted(shops_data["shops"], key=lambda s: s["date"])

    # YEAR FILTER — 2026 only (eliminates 2025-Jun bleed-in bug)
    shops_2026 = [s for s in all_shops if s["date"] >= "2026-01-01"]

    by_month: dict[int, list] = defaultdict(list)
    for s in shops_2026:
        by_month[int(s["date"][5:7])].append(s)

    raw_by_shop: dict = {}
    if PART_JSON.exists():
        raw_by_shop = json.loads(
            PART_JSON.read_text(encoding="utf-8")
        ).get("by_shop", {})

    wb = Workbook()
    wb.remove(wb.active)

    for m in sorted(by_month.keys()):
        month_shops = by_month[m]
        if not month_shops:
            continue
        build_sheet(wb, MONTHS[m], month_shops, raw_by_shop)

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print()

    # Summary for verification
    for m in sorted(by_month.keys()):
        shops = by_month[m]
        hundreds = [s for s in shops if s["score"] == 100.0]
        print(f"  {MONTHS[m]:12s}: {len(shops):2d} shop(s), "
              f"{len(hundreds)} × 100%")

    # Spot-check: June 4 payout
    june_shops = by_month.get(6, [])
    for shop in june_shops:
        if shop["date"] == "2026-06-04":
            jid        = shop["job_id"]
            raw_names  = raw_by_shop.get(jid, [])
            n          = len(raw_names)
            per_person = round(230.0 / n, 2) if n > 0 else 0
            print(f"\n  June 4 check — job {jid}")
            print(f"    Score      : {shop['score']}")
            print(f"    On shift   : {raw_names}  (n={n})")
            print(f"    Per person : ${per_person:.2f}  (target $46.00)")


if __name__ == "__main__":
    build()
