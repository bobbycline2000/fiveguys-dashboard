"""
Build rolling-window hourly curves from Brink Hourly Sales And Labor data.
Median by day-of-week (Mon-Sun) over the last N weeks.

Output: data/case-studies/teamworx-forecast-accuracy/Brink_Rolling_Curves_{store}.xlsx
- Summary tab: count of days per DOW, avg total sales, peak hour
- Mon..Sun tabs: hour-by-hour curve (median), plus all sample days side-by-side
- Curves_JSON tab: machine-readable curve data (for the template builder)

This is the input for the schedule-template product.
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, date
from pathlib import Path
from statistics import median, mean

ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = Path("C:/Users/bobby/OneDrive/BobbyWorkspace")
CASE_STUDY = WORKSPACE / "projects" / "scg" / "case-studies" / "teamworx-forecast-accuracy"
REPO_OUT = ROOT / "data" / "case-studies" / "teamworx-forecast-accuracy"

DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
HOURS_24 = list(range(4, 28))  # 4 AM through 3 AM next day (Brink reporting hours)


def load_brink_data(store: str) -> dict:
    """Returns {date_str: {hour_24: {netSales, guests, orders, laborHrs, laborDollars, laborPct}}}"""
    base = ROOT / "data" / "raw" / "parbrink" / store
    if not base.exists():
        return {}
    out = {}
    for date_dir in sorted(base.iterdir()):
        if not date_dir.is_dir():
            continue
        # Skip non-date directories
        try:
            datetime.strptime(date_dir.name, "%Y-%m-%d")
        except ValueError:
            continue
        # Try JSON first
        j = date_dir / "hourly_sales_labor.json"
        if j.exists():
            data = json.loads(j.read_text())
            rows_list = data.get("rows", [])
            day = {}
            for r in rows_list:
                hour_str = r["hour"].strip()
                h = int(hour_str.split(":")[0])
                if "PM" in hour_str and h != 12:
                    h += 12
                elif "AM" in hour_str and h == 12:
                    h = 0
                day[h] = {k: r.get(k, 0) for k in ("netSales", "guests", "orders", "laborHrs", "laborDollars", "laborPct")}
            out[date_dir.name] = day
            continue
        # Fall back to PDF
        pdf = date_dir / "Hourly Sales And Labor.pdf"
        if pdf.exists():
            try:
                from pypdf import PdfReader
                import re as _re
                reader = PdfReader(str(pdf))
                text = "\n".join(p.extract_text() or "" for p in reader.pages)
                re_row = _re.compile(
                    r"(\d{1,2}:\d{2}\s*[AP]M)\s+\$?([\d,]+\.\d{2})\s+(\d+)\s+\$?[\d,]+\.\d{2}\s+(\d+)\s+\$?[\d,]+\.\d{2}\s+([\d,]+\.\d{2})\s+\$?([\d,]+\.\d{2})\s+([\d,]+\.\d{2})%"
                )
                day = {}
                for m in re_row.finditer(text):
                    hour_str = m.group(1).strip()
                    h = int(hour_str.split(":")[0])
                    if "PM" in hour_str and h != 12:
                        h += 12
                    elif "AM" in hour_str and h == 12:
                        h = 0
                    day[h] = {
                        "netSales": float(m.group(2).replace(",","")),
                        "guests": int(m.group(3)),
                        "orders": int(m.group(4)),
                        "laborHrs": float(m.group(5).replace(",","")),
                        "laborDollars": float(m.group(6).replace(",","")),
                        "laborPct": float(m.group(7).replace(",","")),
                    }
                if day:
                    out[date_dir.name] = day
            except Exception as e:
                print(f"  WARN: failed to parse {pdf}: {e}")
    return out


def hour_label(h: int) -> str:
    h_norm = h % 24
    if h_norm == 0: return "12 AM"
    if h_norm < 12: return f"{h_norm} AM"
    if h_norm == 12: return "12 PM"
    return f"{h_norm - 12} PM"


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--store", default="2065")
    args = p.parse_args()

    brink = load_brink_data(args.store)
    print(f"Loaded {len(brink)} days of Brink data for store {args.store}")
    if not brink:
        return

    # Group by DOW
    by_dow = {dow: {} for dow in DOW_NAMES}  # {dow: {date_str: day_data}}
    for date_str, day in brink.items():
        d = datetime.strptime(date_str, "%Y-%m-%d").date()
        by_dow[DOW_NAMES[d.weekday()]][date_str] = day

    # Compute median + mean curves per DOW
    curves = {}  # {dow: {hour: {netSales_med, netSales_mean, ..., n_samples}}}
    for dow, days in by_dow.items():
        if not days: continue
        hour_curve = {}
        for h in HOURS_24:
            real_h = h if h < 24 else h - 24
            samples = [d.get(real_h, {}).get("netSales", 0) for d in days.values()]
            labor_hrs = [d.get(real_h, {}).get("laborHrs", 0) for d in days.values()]
            labor_dol = [d.get(real_h, {}).get("laborDollars", 0) for d in days.values()]
            guests = [d.get(real_h, {}).get("guests", 0) for d in days.values()]
            hour_curve[h] = {
                "netSales_med": median(samples) if samples else 0,
                "netSales_mean": mean(samples) if samples else 0,
                "labHrs_med": median(labor_hrs) if labor_hrs else 0,
                "labHrs_mean": mean(labor_hrs) if labor_hrs else 0,
                "labDol_mean": mean(labor_dol) if labor_dol else 0,
                "guests_mean": mean(guests) if guests else 0,
                "n": len(samples),
            }
        curves[dow] = {"n_days": len(days), "hours": hour_curve, "dates": sorted(days.keys())}

    # Build Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill(start_color="DA291C", end_color="DA291C", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SECTION_FILL = PatternFill(start_color="FFC72C", end_color="FFC72C", fill_type="solid")

    # Summary tab
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 14
    for c in range(2, 10): ws.column_dimensions[get_column_letter(c)].width = 16

    title = ws.cell(row=1, column=1, value=f"KY-{args.store} — Brink Rolling Hourly Curves (n={len(brink)} days)")
    title.font = Font(bold=True, size=14, color="DA291C")
    ws.merge_cells("A1:I1")

    headers = ["Day", "Sample Days", "Avg Daily $", "Med Daily $", "Peak Hr (med)", "Peak Hr $", "Avg Labor Hrs/day", "Avg Labor $", "Avg Labor %"]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=1+i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    row = 4
    for dow in DOW_NAMES:
        if dow not in curves: continue
        cv = curves[dow]
        days = cv["hours"]
        avg_daily = sum(d["netSales_mean"] for d in days.values())
        med_daily = sum(d["netSales_med"] for d in days.values())
        peak_hr = max(days.items(), key=lambda x: x[1]["netSales_med"])
        avg_lab_hrs = sum(d["labHrs_mean"] for d in days.values())
        avg_lab_dol = sum(d["labDol_mean"] for d in days.values())
        labor_pct = (avg_lab_dol / avg_daily * 100) if avg_daily > 0 else 0
        ws.cell(row=row, column=1, value=dow).font = Font(bold=True)
        ws.cell(row=row, column=2, value=cv["n_days"])
        ws.cell(row=row, column=3, value=avg_daily).number_format = "$#,##0"
        ws.cell(row=row, column=4, value=med_daily).number_format = "$#,##0"
        ws.cell(row=row, column=5, value=hour_label(peak_hr[0]))
        ws.cell(row=row, column=6, value=peak_hr[1]["netSales_med"]).number_format = "$#,##0"
        ws.cell(row=row, column=7, value=round(avg_lab_hrs, 1))
        ws.cell(row=row, column=8, value=avg_lab_dol).number_format = "$#,##0"
        ws.cell(row=row, column=9, value=labor_pct/100).number_format = "0.0%"
        row += 1

    # Per-DOW tabs with hourly detail
    for dow in DOW_NAMES:
        if dow not in curves: continue
        cv = curves[dow]
        ws = wb.create_sheet(dow)
        ws.column_dimensions["A"].width = 12
        for c in range(2, 30): ws.column_dimensions[get_column_letter(c)].width = 11

        title = ws.cell(row=1, column=1, value=f"{dow} — Hourly Curve (n={cv['n_days']} samples: {cv['dates'][0]}..{cv['dates'][-1]})")
        title.font = Font(bold=True, size=12, color="DA291C")
        ws.merge_cells(f"A1:K1")

        # Header row: Hour | Median Sales | Mean Sales | per-date columns
        hdrs = ["Hour", "MEDIAN $", "MEAN $", "Med Lab Hrs", "Mean Lab Hrs", "Lab $/hr (mean)", "Guests Avg"]
        n_per_date = len(cv["dates"])
        for i, h in enumerate(hdrs):
            c = ws.cell(row=3, column=1+i, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
        for i, dt in enumerate(cv["dates"]):
            c = ws.cell(row=3, column=8+i, value=dt)
            c.fill = SECTION_FILL
            c.font = Font(bold=True)

        for ri, h in enumerate(HOURS_24):
            real_h = h if h < 24 else h - 24
            d = cv["hours"][h]
            ws.cell(row=4+ri, column=1, value=hour_label(h)).font = Font(bold=True)
            ws.cell(row=4+ri, column=2, value=round(d["netSales_med"], 2)).number_format = "$#,##0"
            ws.cell(row=4+ri, column=3, value=round(d["netSales_mean"], 2)).number_format = "$#,##0"
            ws.cell(row=4+ri, column=4, value=round(d["labHrs_med"], 2))
            ws.cell(row=4+ri, column=5, value=round(d["labHrs_mean"], 2))
            ws.cell(row=4+ri, column=6, value=round(d["labDol_mean"], 2)).number_format = "$#,##0"
            ws.cell(row=4+ri, column=7, value=round(d["guests_mean"], 1))
            for i, dt in enumerate(cv["dates"]):
                v = brink[dt].get(real_h, {}).get("netSales", 0)
                ws.cell(row=4+ri, column=8+i, value=v).number_format = "$#,##0"
        ws.freeze_panes = "B4"

    # Curves JSON tab
    ws = wb.create_sheet("Curves_JSON")
    ws.column_dimensions["A"].width = 200
    ws.cell(row=1, column=1, value="Machine-readable curves for template builder").font = Font(bold=True)
    ws.cell(row=2, column=1, value=json.dumps(curves, indent=2, default=str)).alignment = Alignment(wrap_text=True, vertical="top")

    # Save
    fname = f"Brink_Rolling_Curves_{args.store}.xlsx"
    REPO_OUT.mkdir(parents=True, exist_ok=True)
    out_repo = REPO_OUT / fname
    try:
        wb.save(out_repo)
        print(f"Wrote {out_repo}")
    except PermissionError:
        from datetime import datetime as _dt2
        ts = _dt2.now().strftime("%Y%m%d-%H%M")
        out_repo = REPO_OUT / f"Brink_Rolling_Curves_{args.store}_{ts}.xlsx"
        wb.save(out_repo)
        print(f"Wrote {out_repo} (timestamped, original was locked)")

    if WORKSPACE.exists():
        try:
            CASE_STUDY.mkdir(parents=True, exist_ok=True)
            wb.save(CASE_STUDY / fname)
            print(f"Also wrote {CASE_STUDY / fname}")
        except PermissionError:
            print(f"Workspace path locked, skipped")

    # Also save curves as standalone JSON for the template builder to consume
    json_out = REPO_OUT / f"Brink_Rolling_Curves_{args.store}.json"
    json_out.write_text(json.dumps(curves, indent=2, default=str))
    print(f"Wrote {json_out}")


if __name__ == "__main__":
    main()
