"""Build employee directory Excel file from the handwritten list."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

EMPLOYEES = [
    ("Alen",      "(407) 334-3088", "Active"),
    ("Ash",       "(502) 554-4833", "Active"),
    ("Autumn",    "(502) 656-2133", "Active"),
    ("Bri",       "(502) 202-4126", "Active"),
    ("Divan",     "(502) 821-9406", "Active"),
    ("Francisco", "(407) 456-1445", "Active"),
    ("Grace",     "(502) 527-1477", "Active"),
    ("Jada",      "(502) 821-5962", "Active"),
    ("Jeremiah",  "(502) 498-9401", "Active"),
    ("Kable",     "(502) 606-3088", "Active"),
    ("Kasey",     "(502) 884-2273", "Active"),
    ("Kayla",     "(502) 531-8704", "Active"),
    ("Kenzie",    "(502) 650-2273", "Active"),
    ("Lidy",      "(502) 381-0680", "Active"),
    ("Madison",   "(502) 438-7200", "Active"),
    ("Maylin",    "(502) 803-1973", "Active"),
    ("Mike",      "(502) 281-8177", "Active"),
    ("Nathan",    "(502) 693-8371", "Active"),
    ("Richard",   "(270) 874-0655", "Active"),
    ("Samuel",    "(502) 724-2384", "Active"),
    ("Vicki",     "(502) 701-9597", "Active"),
]


def build(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ["Name", "Phone Number", "Status"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row in EMPLOYEES:
        ws.append(row)

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
        if ws.cell(row=r, column=3).value and "Crossed out" in ws.cell(row=r, column=3).value:
            grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = grey
                ws.cell(row=r, column=c).font = Font(italic=True, color="777777")

    widths = {1: 16, 2: 20, 3: 38}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    wb.save(path)
    print(f"Wrote {path} with {len(EMPLOYEES)} rows")


def build_pdf(path: str) -> None:
    doc = SimpleDocTemplate(
        path,
        pagesize=LETTER,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        title="Employee Directory — Five Guys 2065",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=16, textColor=colors.HexColor("#C8102E"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "subtitle", parent=styles["Normal"],
        fontName="Helvetica", fontSize=10, textColor=colors.HexColor("#666666"),
        spaceAfter=14,
    )

    story = [
        Paragraph("Employee Directory", title_style),
        Paragraph("Five Guys — Store 2065, Dixie Highway, Louisville KY", subtitle_style),
    ]

    data = [["Name", "Phone Number"]]
    for name, phone, _status in EMPLOYEES:
        data.append([name, phone])

    table = Table(data, colWidths=[2.0 * inch, 2.5 * inch], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C8102E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.25 * inch))
    story.append(Paragraph(
        f"{len(EMPLOYEES)} employees",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=9,
                       textColor=colors.HexColor("#888888"), alignment=2),
    ))

    doc.build(story)
    print(f"Wrote {path} with {len(EMPLOYEES)} rows")


if __name__ == "__main__":
    build("employee_directory.xlsx")
    build_pdf("employee_directory.pdf")
