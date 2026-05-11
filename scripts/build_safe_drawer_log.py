"""Build the monthly Safe & Drawer Count Log (xlsx + pdf).

Five Guys Store 2065 — Louisville KY.

Banks:
  - Safe bank: $700.00
  - Drawer 1:  $200.00
  - Drawer 2:  $200.00

The log is a daily template managers fill out:
  - Opening counts for safe + both drawers
  - Closing counts
  - Deposit bag # + amount
  - CrunchTime over/short for the day (filled at EOD ~4 AM next day)
  - Manager initials + notes

Run:  python scripts/build_safe_drawer_log.py
Out:  safe_drawer_log.xlsx, safe_drawer_log.pdf in repo root.
"""
from __future__ import annotations

import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

SAFE_BANK = 700.00
DRAWER_BANK = 200.00
NUM_DRAWERS = 2

HEADERS = [
    "Date",
    "Mgr Initials",
    f"Safe Open (${SAFE_BANK:.0f})",
    f"Drawer 1 Open (${DRAWER_BANK:.0f})",
    f"Drawer 2 Open (${DRAWER_BANK:.0f})",
    "Drawer 1 Close",
    "Drawer 2 Close",
    "Deposit Bag #",
    "Deposit Amount",
    "CT Over/Short",
    "Notes",
]


def _month_dates(year: int, month: int) -> list[date]:
    last = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, last + 1)]


def build_xlsx(path: str, year: int, month: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1)
    title.value = (
        f"Five Guys 2065 — Safe & Drawer Count Log — "
        f"{calendar.month_name[month]} {year}"
    )
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Subtitle: bank reminders
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    sub = ws.cell(row=2, column=1)
    sub.value = (
        f"Safe bank: ${SAFE_BANK:,.2f}   |   Drawers: ${DRAWER_BANK:,.2f} each "
        f"(2 drawers)   |   Deposit nightly + enter in CrunchTime"
    )
    sub.font = Font(italic=True, color="555555", size=10)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 30

    # Date rows
    dates = _month_dates(year, month)
    row_start = 4
    for i, d in enumerate(dates):
        r = row_start + i
        weekend = d.weekday() >= 5
        row_fill = PatternFill(
            start_color="FFF9E6" if weekend else "FFFFFF",
            end_color="FFF9E6" if weekend else "FFFFFF",
            fill_type="solid",
        )
        ws.cell(row=r, column=1, value=d.strftime("%a %m/%d"))
        # Pre-fill the expected bank amounts
        ws.cell(row=r, column=3, value=SAFE_BANK)
        ws.cell(row=r, column=4, value=DRAWER_BANK)
        ws.cell(row=r, column=5, value=DRAWER_BANK)
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if c in (3, 4, 5, 6, 7, 9, 10):
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Column widths
    widths = {1: 13, 2: 12, 3: 16, 4: 18, 5: 18, 6: 16, 7: 16, 8: 14, 9: 16, 10: 16, 11: 32}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"

    # Totals row at the bottom
    total_row = row_start + len(dates) + 1
    ws.cell(row=total_row, column=1, value="MONTH TOTALS").font = Font(bold=True)
    for col_letter in ("I", "J"):
        c = ws[f"{col_letter}{total_row}"]
        c.value = f"=SUM({col_letter}{row_start}:{col_letter}{row_start + len(dates) - 1})"
        c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True)

    wb.save(path)
    print(f"Wrote {path}")


def build_pdf(path: str, year: int, month: int) -> None:
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(LETTER),
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        title=f"Safe & Drawer Count Log — {calendar.month_name[month]} {year}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        fontName="Helvetica-Bold", fontSize=14, textColor=colors.HexColor("#C8102E"),
        spaceAfter=2, alignment=1,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#555555"),
        spaceAfter=10, alignment=1,
    )

    story = [
        Paragraph(
            f"Five Guys 2065 — Safe &amp; Drawer Count Log — "
            f"{calendar.month_name[month]} {year}", title_style,
        ),
        Paragraph(
            f"Safe bank: ${SAFE_BANK:,.2f} &nbsp;|&nbsp; "
            f"Drawers: ${DRAWER_BANK:,.2f} each (2) &nbsp;|&nbsp; "
            f"Deposit nightly + enter in CrunchTime",
            sub_style,
        ),
    ]

    data = [HEADERS]
    for d in _month_dates(year, month):
        data.append([
            d.strftime("%a %m/%d"),
            "",
            f"${SAFE_BANK:.0f}",
            f"${DRAWER_BANK:.0f}",
            f"${DRAWER_BANK:.0f}",
            "", "", "", "", "", "",
        ])

    col_widths = [0.85, 0.65, 0.95, 1.0, 1.0, 0.9, 0.9, 0.8, 0.95, 0.95, 1.6]
    col_widths = [w * inch for w in col_widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2A44")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#BBBBBB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F8F8")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        "Manager signs each day. Discrepancies &gt; $5: note cause in Notes. "
        "Over/Short comes from CrunchTime Total Cash Over/Short after EOD (~4 AM).",
        ParagraphStyle("foot", parent=styles["Normal"], fontSize=8,
                       textColor=colors.HexColor("#888888"), alignment=1),
    ))
    doc.build(story)
    print(f"Wrote {path}")


if __name__ == "__main__":
    today = date.today()
    build_xlsx("safe_drawer_log.xlsx", today.year, today.month)
    build_pdf("safe_drawer_log.pdf", today.year, today.month)
